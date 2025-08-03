from django.contrib import messages
from django.http import JsonResponse
from django.shortcuts import render, redirect
from . import models, forms
from course.models import Course
from openpyxl import load_workbook
from django.contrib.auth.decorators import login_required

def get_user_school(request):
    if request.user.is_superuser:
        return None
    return request.user.userprofile.school

def index(request):
    school = get_user_school(request)
    students = models.Student.objects.all() if school is None else models.Student.objects.filter(school=school)
    return render(request, 'student/index.html', {'students': students})

def create(request):
    form = forms.CreateStudentForm(user=request.user)
    return render(request, 'student/create.html', {'form': form})

def store(request):
    if request.method == 'POST':
        form = forms.CreateStudentForm(request.POST, user=request.user)
        if form.is_valid():
            student = form.save(commit=False)
            if not request.user.is_superuser:
                student.school = request.user.userprofile.school
            student.save()
            messages.success(request, 'Student created successfully')
            return redirect('student.index')
        else:
            return render(request, 'student/create.html', {'form': form})
    return redirect('student.create')

def edit(request, sid):
    school = get_user_school(request)
    try:
        student = models.Student.objects.get(id=sid) if school is None else models.Student.objects.get(id=sid, school=school)
        form = forms.CreateStudentForm(instance=student, user=request.user)
        return render(request, 'student/edit.html', {'form': form})
    except models.Student.DoesNotExist:
        return redirect('student.index')

def update(request, sid):
    school = get_user_school(request)
    if request.method == 'POST':
        try:
            student = models.Student.objects.get(id=sid) if school is None else models.Student.objects.get(id=sid, school=school)
            form = forms.CreateStudentForm(request.POST, instance=student, user=request.user)
            if form.is_valid():
                form.save()
                messages.success(request, 'Student updated successfully')
                return redirect('student.index')
            else:
                return render(request, 'student/edit.html', {'form': form})
        except models.Student.DoesNotExist:
            return redirect('student.index')
    return redirect('student.index')

def delete(request, sid):
    school = get_user_school(request)
    if request.method == 'POST':
        try:
            student = models.Student.objects.get(id=sid) if school is None else models.Student.objects.get(id=sid, school=school)
            student.delete()
            messages.success(request, 'Student deleted successfully')
        except models.Student.DoesNotExist:
            pass
    return redirect('student.index')

def get_course_fee(request):
    course_id = request.GET.get('course_id')
    if course_id:
        try:
            course = Course.objects.get(pk=course_id)
            return JsonResponse({'fee': course.total_amount})
        except Course.DoesNotExist:
            return JsonResponse({'fee': 0})
    return JsonResponse({'fee': 0})

@login_required
def import_students(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        file = request.FILES["excel_file"]
        try:
            wb = load_workbook(file)
        except Exception:
            messages.error(request, "Failed to open the Excel file. Please ensure it is a valid .xlsx file.")
            return redirect('student.import')

        sheet = wb.active
        created_count = 0
        skipped_rows = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            name, email, contact, address, course_name = row[:5]  # adapt based on your sheet
            if not name:
                skipped_rows.append(idx)
                continue

            course = None
            if course_name:
                course = Course.objects.filter(name__iexact=course_name.strip()).first()

            models.Student.objects.create(
                name=name.strip(),
                email=email or "",
                contact=contact or "",
                address=address or "",
                course=course,
                school=request.user.userprofile.school if not request.user.is_superuser else None,
            )
            created_count += 1

        if created_count:
            messages.success(request, f"{created_count} students imported successfully.")
        if skipped_rows:
            messages.warning(request, f"Skipped rows missing name: {skipped_rows}")

        return redirect('student.index')

    return render(request, "student/import.html")
